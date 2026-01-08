import argparse
import openpyxl
import docx2pdf
from docx import Document
from pathlib import Path


def scan_excel_headers(file):
    excel_sheet = openpyxl.load_workbook(file).active
    headers = []

    for i in range(1, excel_sheet.max_column + 1):
        headers.append("<" + excel_sheet.cell(row=1, column=i).value + ">")

    return headers


def get_excel_total_row_values(file):
    excel_sheet = openpyxl.load_workbook(file).active
    return excel_sheet.max_row - 1 # -1 to remove header from total rows


def scan_excel_row(file, headers, row):
    excel_sheet = openpyxl.load_workbook(file).active
    values_dict = {}

    for i in range(1, excel_sheet.max_column + 1):
        values_dict[headers[i-1]] = excel_sheet.cell(row=row, column=i).value
    
    return values_dict


def docx_replace(file, values, output_filename, directory):
    document = Document(file)

    for k, v in values.items():
        for paragraph in document.paragraphs:
            if k in paragraph.text:
                paragraph.text = paragraph.text.replace(k, str(v))
    
    output_file = Path(directory) / f"{output_filename}.docx"
    document.save(output_file)
    return output_file


def docx_pdf_export(docx_file, output_filename, directory):
    output_file = Path(directory) / f"{output_filename}.pdf"
    docx2pdf.convert(docx_file,output_file)


# Parse commandline arguments and assign them to pertinent variables
parser = argparse.ArgumentParser(description="Applies values in an Excel file to a Word Template and Export them to PDF")
parser.add_argument("doc_template", help=".docx template for Excel values to substitute to")
parser.add_argument("excel_sheet", help=".xlsx file containing columns to represent document template variable and rows to represent the different values associated with the template variable")
parser.add_argument("suffix", help="the suffix added to the filename of the outputs")
parser.add_argument("-o", "--output_directory", help="Optional directory for output. Default is the current directory.")
args = parser.parse_args()

doc_template = args.doc_template
xl_sheet = args.excel_sheet
suffix = args.suffix
directory = "."
if args.output_directory:
    directory = args.output_directory

xl_headers = scan_excel_headers(xl_sheet)
xl_total_values = get_excel_total_row_values(xl_sheet)
for i in range(2, xl_total_values + 2): # 2 to skip header
    current_xl_values = scan_excel_row(xl_sheet, xl_headers, i)
    output_filename = current_xl_values[xl_headers[0]] + suffix
    output_document = docx_replace(doc_template, current_xl_values, output_filename, directory)
    docx_pdf_export(output_document, output_filename, directory)


# Tests
#
#
# TEST CMDLINE ARGUMENTS
# print(f"Document: {doc_template}\nExcel Sheet: {xl_sheet}\nSuffix: {suffix}\nDirectory: {directory}") # Test commandline arguments
#
#
# TEST SCAN HEADERS
# for i in xl_headers:
#   print(i)
#
#
# TEST GET TOTAL VALUES
# print(xl_total_values)
#
#
# TEST SCAN EXCEL ROW
# for i in range(2, xl_total_values + 2): # 2 to skip header
#   current_xl_values = scan_excel_row(xl_sheet, xl_headers, i)
#   for k, v in current_xl_values.items():
#       print(k,v)
#   print("")
